# -*- coding: utf-8 -*-
"""
Created on Wed Aug 14 14:03:03 2024

@author: roone
"""

import os
import pandas as pd
import openpyxl

folder_path = r"C:\Users\roone\Downloads\OneDrive_2024-08-09\VMM Lifetime Logs"
columns = ["file name", "first day", "last day", "Inspection Interval Shock", "Cumulative Hours Shock", "Hours Shock", "Max Lateral Vib"]
results_df = pd.DataFrame(columns=columns)

for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx"):
        file_path = os.path.join(folder_path, filename)
        print(f"Processing file: {filename}")

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active

            # New columns data
            inspection_interval_shock = [sheet.cell(row=12, column=col).value for col in range(66, 120)]  # Columns BN to CU
            cumulative_hours_shock = [sheet.cell(row=13, column=col).value for col in range(66, 120)]  # Columns BN to CU
            hours_shock = [sheet.cell(row=14, column=col).value for col in range(66, 120)]  # Columns BN to CU
            max_lateral_vib = [sheet.cell(row=15, column=col).value for col in range(66, 120)]  # Columns BN to CU

            first_day = sheet['I4'].value  
            last_day = sheet['I5'].value   
            print(f"First day: {first_day}, Last day: {last_day}")
            print(f"Inspection Interval Shock: {inspection_interval_shock[:5]}...")
            print(f"Cumulative Hours Shock: {cumulative_hours_shock[:5]}...")
            print(f"Hours Shock: {hours_shock[:5]}...")
            print(f"Max Lateral Vib: {max_lateral_vib[:5]}...")

            file_df = pd.DataFrame({
                "file name": [filename] * len(max_lateral_vib),
                "first day": [first_day] * len(max_lateral_vib),
                "last day": [last_day] * len(max_lateral_vib),
                "Inspection Interval Shock": inspection_interval_shock,
                "Cumulative Hours Shock": cumulative_hours_shock,
                "Hours Shock": hours_shock,
                "Max Lateral Vib": max_lateral_vib
            })
            results_df = pd.concat([results_df, file_df], ignore_index=True)

        except Exception as e:
            print(f"Skipping file {filename} due to error: {e}")

output_path = os.path.join(folder_path, "Aggregated_Results2.xlsx")
results_df.to_excel(output_path, index=False)
print(f"Results saved to {output_path}")
